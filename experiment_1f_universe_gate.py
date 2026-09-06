"""Experiment #1F Gate A — data-integration gate for the approved 298 universe.

WHY THIS FILE EXISTS
--------------------
Experiment #1/#1E ran on the 45-ticker conglomerate watchlist in neobdm.db. The
approved research universe is a different, broader set (298 names). Only 30 of
those 298 have any row in neobdm.db, so #1F cannot source prices from the
production database. The data does exist -- in the ohlc/broker_daily parquet
harvest -- but that harvest has never been through the validation stack that
Experiment #1E spent an entire review cycle hardening.

Running #1F straight off raw parquet would therefore discard the executable
contract. This module is the gate that prevents that: it puts the harvest
through the SAME checks, reusing price_audit's real functions rather than
reimplementing them, and refuses to hand Gate B a panel that has not passed.

THE DETECTOR SEES THE WHOLE MARKET, THE PANEL SEES THE UNIVERSE
---------------------------------------------------------------
cross_ticker_dup is a CROSS-SECTIONAL detector: it finds one stock's OHLCV
copied onto another, which is what a chart that had not re-rendered yet
produces. Subsetting to the approved names before running it would hide exactly
the contamination it exists to catch -- an approved ticker cloned against a
ticker OUTSIDE the universe looks perfectly unique once the other half of the
pair has been filtered away. So detect() runs on the FULL frozen harvest, and
only afterwards are the suspect keys and the model panel narrowed to the
approved names. The trading calendar likewise comes from the full harvest.

FROZEN, NOT REFRESHED
---------------------
The harvest is deliberately NOT re-run before validation. The harvest/backfill
path writes price history with INSERT OR REPLACE across the whole chart range,
so a refresh restates months of history rather than appending to it -- a
point-in-time hazard (today's revision silently replacing what was observable
then) and a determinism hazard (a frozen experiment whose inputs move). The
reviewed snapshot is pinned by a SHA-256 input manifest that must already exist:
a missing manifest is a failure, never an invitation to adopt whatever bytes
happen to be on disk.

WHAT RAISES AND WHAT DOES NOT
-----------------------------
Structural invariants that no downstream code could survive -- duplicate keys,
impossible calendar dates, non-finite prices, cross-ticker contamination
surviving quarantine, a label leaking across a ticker boundary -- RAISE. So do
changed input fingerprints, changed coverage, and a changed universe.

An invalid OPEN anchor does NOT raise. That is a per-row data-quality fact the
executable contract already handles by withholding the label: the row survives,
its fwd_oo_* go NaN. Deleting such rows would be a repair, and this module
repairs nothing.

WHAT THIS MODULE DOES NOT DO
----------------------------
It does not train, score, rank, or simulate anything -- that is Gate B. It does
not write to neobdm.db or touch any Experiment #1/#1E artifact. It does not
backfill or fabricate a single price.
"""

import argparse
import datetime as _dt
import hashlib
import json
import os
import re

import numpy as np
import pandas as pd

from price_audit import (
    OHLCV, add_forward_returns, add_lagged_returns, detect, _open_anchor_valid,
)

HERE = os.path.dirname(os.path.abspath(__file__))

#: The approved universe as supplied by the user. Read-only: this module never
#: writes to it. Once frozen into FROZEN_UNIVERSE_JSON that artifact becomes the
#: reproducible source of record, so #1F does not depend on a Desktop path
#: staying put -- see resolve_universe().
UNIVERSE_XLSX = os.path.join(
    r"C:\Users\jason\Desktop\Inventory + Broker Summary ML\Extract Data dulu",
    "Stock Universe Set.xlsx",
)
UNIVERSE_SHEET = "Universe for ML"
UNIVERSE_HEADER = "symbol"
FROZEN_UNIVERSE_JSON = os.path.join(HERE, "experiment_1f_universe.json")
INPUT_MANIFEST_JSON = os.path.join(HERE, "experiment_1f_input_manifest.json")

OHLC_PARQUET = os.path.join(HERE, "ohlc.parquet")
BROKER_PARQUET = os.path.join(HERE, "broker_daily.parquet")

#: A ticker is four uppercase letters. Enforced as a real pattern, not implied
#: by a digest: duplicates, blanks and type corruption can all leave the
#: DISTINCT set -- and therefore the digest -- completely unchanged.
TICKER_PATTERN = re.compile(r"^[A-Z]{4}$")
#: Shape check only. Shape is NOT validity: 2026-99-77 matches this pattern and
#: is not a date, so every date check pairs it with strict parsing.
DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")
DATE_FORMAT = "%Y-%m-%d"

#: Hard expectations for the approved freeze. Any deviation stops the gate.
EXPECTED_UNIVERSE_SIZE = 298
ACCEPTED_UNIVERSE_DIGEST = "7a4ce50ffb5a7ebe"
EXPECTED_PRICE_COVERED = 297
EXPECTED_PRICE_MISSING = ("WIKA",)
#: Broker coverage is frozen independently of price coverage: a broker harvest
#: that quietly covered fewer names would silently shrink variants B/C/D without
#: shrinking A, making the variants incomparable.
EXPECTED_BROKER_COVERED = 297
EXPECTED_BROKER_MISSING = ("WIKA",)

#: Where the harvest's broker values can be checked against independently
#: observed data. neobdm.db holds live-scraped rows (bval IS NOT NULL) from this
#: date onward; earlier harvest rows have nothing to check against.
#:
#: This window establishes VALUE FIDELITY, NOT point-in-time availability. See
#: broker_provenance() for why the distinction matters.
VALUE_RECONCILED_FROM = "2026-07-05"

#: #1E's feature code (_historical_net_lots) reads netval/bval/sval in BILLIONS
#: of rupiah; the harvest stores plain rupiah. One conversion, applied once, and
#: proven against the live rows rather than assumed.
RUPIAH_PER_BILLION = 1e9

#: Lots are 100 shares on IDX. Used only to recover the average execution price
#: implied by (value, lots).
SHARES_PER_LOT = 100.0

#: Rupiah amounts are whole numbers in practice, and float64 carries ~15-16
#: significant digits, so a value near 1e12 still resolves far below 1 rupiah.
#: A 1-rupiah tolerance is therefore below the smallest meaningful unit while
#: absorbing any parquet round-trip representation error.
NVAL_TOLERANCE_RUPIAH = 1.0
#: Lots are integers; this only absorbs float representation noise.
NLOT_TOLERANCE = 1e-6

REQUIRED_BROKER_NUMERIC = ("nlot", "blot", "slot", "nval", "bval", "sval")

HORIZONS = (1, 2, 3, 4, 5)
LAGS = (1, 3, 5, 10, 20)


class GateFailure(AssertionError):
    """Raised when an input violates a frozen expectation of the gate."""


def _digest(values):
    """SHA-256[:16] over a '|'-joined payload.

    Same construction as ml_v2_experiment_1.prediction_digest so digests across
    experiments are read the same way.
    """
    return hashlib.sha256("|".join(values).encode("utf-8")).hexdigest()[:16]


def frame_digest(df, columns=None):
    """Content digest of a frame, float-formatted at .12g.

    .12g is what #1E pins its predictions with: stable across platforms, but
    sensitive to any real numeric change.
    """
    cols = list(columns or df.columns)
    values = ["|".join(cols)]
    for col in cols:
        series = df[col]
        if pd.api.types.is_float_dtype(series):
            values.extend("" if pd.isna(v) else f"{v:.12g}" for v in series.to_numpy())
        else:
            values.extend("" if pd.isna(v) else str(v) for v in series.to_numpy())
    return _digest(values)


def invalid_date_mask(series):
    """Rows whose date is missing, misshapen, or not a real calendar date.

    Shape and validity are different questions. '2026-99-77' satisfies
    ^\\d{4}-\\d{2}-\\d{2}$ and is not a date, so the regex is paired with strict
    parsing; coerce turns anything unparseable into NaT rather than guessing at
    a different format.
    """
    shape_ok = series.astype(str).str.match(DATE_PATTERN)
    parsed = pd.to_datetime(series, format=DATE_FORMAT, errors="coerce")
    return series.isna() | ~shape_ok.fillna(False) | parsed.isna()


def file_fingerprint(path):
    """Full SHA-256 + size of a raw input snapshot."""
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    stat = os.stat(path)
    return {
        "path": os.path.basename(path),
        "sha256": digest.hexdigest(),
        "size_bytes": stat.st_size,
        "mtime_utc": _dt.datetime.utcfromtimestamp(stat.st_mtime).isoformat() + "Z",
    }


def verify_input_manifest(fingerprints, path=INPUT_MANIFEST_JSON, establish=False):
    """Pin the reviewed raw snapshot so a refreshed harvest cannot slip in.

    A MISSING manifest is a failure, not an invitation. Auto-establishing one
    would reopen the exact hole this closes: delete the manifest, refresh the
    parquet, and whatever bytes happen to be on disk quietly become the
    "reviewed" snapshot. Establishing or re-establishing is therefore an
    explicit, reviewed action (--establish-manifest).

    `mtime` is recorded but never compared: touching a file does not change its
    content.
    """
    current = {fp["path"]: {"sha256": fp["sha256"], "size_bytes": fp["size_bytes"]}
               for fp in fingerprints}

    if establish:
        payload = {
            "established_at_utc": _dt.datetime.utcnow().isoformat() + "Z",
            "inputs": current,
        }
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, sort_keys=True)
            handle.write("\n")
        return {"status": "ESTABLISHED (explicit, reviewed)",
                "manifest_path": path, "inputs": current}

    if not os.path.exists(path):
        raise GateFailure(
            f"frozen input manifest {os.path.basename(path)} is missing. Gate A "
            "will not adopt whatever parquet bytes are on disk as the reviewed "
            "snapshot -- that is how a silently refreshed harvest becomes the "
            "baseline. Review the hashes, then re-run with --establish-manifest.")

    with open(path, encoding="utf-8") as handle:
        stored = json.load(handle)["inputs"]

    failures = []
    for name, values in current.items():
        if name not in stored:
            failures.append(f"{name} is absent from the reviewed manifest")
            continue
        if stored[name]["sha256"] != values["sha256"]:
            failures.append(
                f"{name} sha256 {values['sha256'][:16]}... != reviewed "
                f"{stored[name]['sha256'][:16]}... (the snapshot changed)")
        if stored[name]["size_bytes"] != values["size_bytes"]:
            failures.append(
                f"{name} size {values['size_bytes']} != reviewed "
                f"{stored[name]['size_bytes']}")
    for name in stored:
        if name not in current:
            failures.append(f"{name} is in the reviewed manifest but was not loaded")
    if failures:
        raise GateFailure(
            "frozen input snapshot no longer matches the reviewed manifest. A new "
            "hash is never silently accepted -- re-establish deliberately with "
            "--establish-manifest after review:\n  - " + "\n  - ".join(failures))
    return {"status": "verified", "manifest_path": path, "inputs": current}


# ── universe ingestion ────────────────────────

def validate_universe_cells(cells, source_path=UNIVERSE_XLSX, sheet=UNIVERSE_SHEET):
    """HARD-VALIDATE a column of universe cells. Raises GateFailure on any defect.

    The audit is not ceremonial. A spreadsheet is exactly where a legitimate
    ticker turns into something else: IDX lists a real stock called TRUE, which
    Excel will happily store as the boolean TRUE, and a ticker that looks
    numeric or date-like can be retyped the same way.

    The checks do NOT rely on the digest alone, because the digest covers the
    sorted DISTINCT set: a duplicated row, a blank cell, or a coerced type can
    leave that set -- and therefore the digest -- byte-identical while the file
    has genuinely changed.

    Split out from the workbook reader so the validation is testable without a
    spreadsheet on disk.
    """
    header, body = (cells[0] if cells else None), list(cells[1:])
    non_string = [{"row": i + 2, "value": repr(v), "type": type(v).__name__}
                  for i, v in enumerate(body)
                  if v is not None and not isinstance(v, str)]
    blanks = [i + 2 for i, v in enumerate(body)
              if v is None or (isinstance(v, str) and not v.strip())]
    strings = [v for v in body if isinstance(v, str) and v.strip()]
    malformed = [{"row": i + 2, "value": repr(v)} for i, v in enumerate(body)
                 if isinstance(v, str) and v.strip() and not TICKER_PATTERN.match(v)]

    counts = pd.Series(strings, dtype=object).value_counts() if strings else pd.Series(dtype=int)
    duplicates = sorted(counts[counts > 1].index.tolist())
    tickers = sorted(set(strings))
    digest = _digest(tickers)

    audit = {
        "source_path": source_path,
        "sheet": sheet,
        "header_cell": header,
        "rows_after_header": len(body),
        "non_string_cells": non_string,
        "blank_cells": blanks,
        "malformed_tickers": malformed,
        "duplicates": duplicates,
        "distinct_tickers": len(tickers),
        "universe_digest": digest,
    }

    failures = []
    if header != UNIVERSE_HEADER:
        failures.append(f"header cell is {header!r}, expected {UNIVERSE_HEADER!r}")
    if non_string:
        failures.append(f"{len(non_string)} non-string cell(s) (Excel type coercion): {non_string[:5]}")
    if blanks:
        failures.append(f"{len(blanks)} blank cell(s) at rows {blanks[:10]}")
    if malformed:
        failures.append(f"{len(malformed)} malformed ticker(s) failing ^[A-Z]{{4}}$: {malformed[:5]}")
    if duplicates:
        failures.append(f"{len(duplicates)} duplicate ticker(s): {duplicates[:10]}")
    if len(tickers) != EXPECTED_UNIVERSE_SIZE:
        failures.append(f"{len(tickers)} distinct tickers, expected {EXPECTED_UNIVERSE_SIZE}")
    if digest != ACCEPTED_UNIVERSE_DIGEST:
        failures.append(f"digest {digest} != accepted {ACCEPTED_UNIVERSE_DIGEST}")
    if failures:
        raise GateFailure(
            "approved universe failed validation - re-freezing is a reviewed "
            "decision, never automatic:\n  - " + "\n  - ".join(failures))

    return tickers, audit


def read_universe_xlsx(path=UNIVERSE_XLSX, sheet=UNIVERSE_SHEET):
    """Read the approved workbook column and hand it to validate_universe_cells."""
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        workbook.close()
        raise GateFailure(
            f"sheet {sheet!r} not found in {path} (sheets: {workbook.sheetnames})")
    cells = [row[0].value for row in workbook[sheet].iter_rows(min_col=1, max_col=1)]
    workbook.close()
    return validate_universe_cells(cells, source_path=path, sheet=sheet)


def write_frozen_universe(tickers, audit, path=FROZEN_UNIVERSE_JSON):
    """Persist the validated universe as the reproducible source of record.

    `frozen_at_utc` is provenance metadata only. It is deliberately NOT part of
    the universe digest, so re-freezing identical content yields an identical
    digest and the artifact stays deterministic.
    """
    payload = {
        "universe_digest": audit["universe_digest"],
        "n_tickers": len(tickers),
        "source_path": audit["source_path"],
        "source_sheet": audit["sheet"],
        "frozen_at_utc": _dt.datetime.utcnow().isoformat() + "Z",
        "tickers": tickers,
    }
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, sort_keys=True)
        handle.write("\n")
    return payload


def load_frozen_universe(path=FROZEN_UNIVERSE_JSON):
    """Load and re-verify the frozen artifact, independent of the workbook."""
    with open(path, encoding="utf-8") as handle:
        payload = json.load(handle)
    tickers = payload["tickers"]

    failures = []
    if sorted(set(tickers)) != sorted(tickers):
        failures.append("frozen artifact contains duplicate tickers")
    bad = [t for t in tickers if not isinstance(t, str) or not TICKER_PATTERN.match(t)]
    if bad:
        failures.append(f"frozen artifact contains malformed tickers: {bad[:5]}")
    if len(tickers) != EXPECTED_UNIVERSE_SIZE:
        failures.append(f"{len(tickers)} tickers, expected {EXPECTED_UNIVERSE_SIZE}")
    recomputed = _digest(sorted(tickers))
    if recomputed != ACCEPTED_UNIVERSE_DIGEST:
        failures.append(f"recomputed digest {recomputed} != accepted {ACCEPTED_UNIVERSE_DIGEST}")
    if payload.get("universe_digest") != recomputed:
        failures.append(
            f"stored digest {payload.get('universe_digest')} != recomputed {recomputed}")
    if failures:
        raise GateFailure("frozen universe artifact failed validation:\n  - "
                          + "\n  - ".join(failures))

    audit = {
        "source_path": payload.get("source_path"),
        "sheet": payload.get("source_sheet"),
        "header_cell": UNIVERSE_HEADER,
        "rows_after_header": None,
        "non_string_cells": [],
        "blank_cells": [],
        "malformed_tickers": [],
        "duplicates": [],
        "distinct_tickers": len(tickers),
        "universe_digest": recomputed,
        "loaded_from": "frozen artifact",
        "frozen_at_utc": payload.get("frozen_at_utc"),
    }
    return sorted(tickers), audit


def resolve_universe(xlsx_path=UNIVERSE_XLSX, frozen_path=FROZEN_UNIVERSE_JSON,
                     refreeze=False):
    """Prefer the frozen artifact; fall back to the workbook only to create it.

    Gate B must never reread a mutable Desktop workbook, so once the freeze
    exists it is authoritative. `refreeze=True` re-reads the workbook
    deliberately -- and still hard-fails if its content no longer matches the
    accepted freeze, so a changed workbook can never be picked up silently.
    """
    if os.path.exists(frozen_path) and not refreeze:
        return load_frozen_universe(frozen_path)
    tickers, audit = read_universe_xlsx(xlsx_path)
    payload = write_frozen_universe(tickers, audit, frozen_path)
    audit["loaded_from"] = "workbook (frozen now)"
    audit["frozen_at_utc"] = payload["frozen_at_utc"]
    return tickers, audit


# ── raw source integrity ──────────────────────

def audit_raw_ohlc(frame, calendar, scope="full harvest"):
    """Structural audit of the RAW price harvest, BEFORE detect() runs.

    Runs on the FULL harvest, because detect() runs on the full harvest: its
    cross-sectional duplicate test is only meaningful over the whole
    cross-section, and it assumes the frame is already structurally sound --
    unique keys, real dates, finite numbers. If that assumption is wrong its
    output is meaningless, so those properties are established first, here.

    Non-positive OPEN is reported but does NOT fail: an unusable open is exactly
    what the open-anchor contract exists to catch, and it withholds the label
    rather than deleting the row. Non-positive high/low/close DO fail -- nothing
    downstream can form a return from them and no contract absorbs it.
    """
    report = {"scope": scope, "rows": int(len(frame))}
    failures = []

    report["duplicate_date_ticker"] = int(frame.duplicated(["date", "ticker"]).sum())
    if report["duplicate_date_ticker"]:
        failures.append(f"{report['duplicate_date_ticker']} duplicate (date,ticker) rows")

    bad_date = invalid_date_mask(frame["date"])
    report["invalid_dates"] = int(bad_date.sum())
    if report["invalid_dates"]:
        examples = frame.loc[bad_date, "date"].astype(str).unique()[:5].tolist()
        failures.append(
            f"{report['invalid_dates']} missing/unparseable/impossible dates, "
            f"e.g. {examples}")

    missing_ticker = frame["ticker"].isna() | (frame["ticker"].astype(str).str.strip() == "")
    report["missing_ticker"] = int(missing_ticker.sum())
    if report["missing_ticker"]:
        failures.append(f"{report['missing_ticker']} rows with a missing ticker")

    non_finite = {}
    for col in OHLCV:
        count = int((~np.isfinite(frame[col].to_numpy(dtype=float))).sum())
        non_finite[col] = count
        if count:
            failures.append(f"{count} non-finite values in {col}")
    report["non_finite"] = non_finite

    nonpositive = {col: int((frame[col] <= 0).sum()) for col in ("open", "high", "low", "close")}
    report["nonpositive"] = nonpositive
    for col in ("high", "low", "close"):
        if nonpositive[col]:
            failures.append(f"{nonpositive[col]} non-positive {col}")
    report["nonpositive_open_note"] = (
        "non-positive open is reported, not fatal: the open-anchor contract "
        "withholds the label for such a row rather than deleting it")

    calendar_series = pd.Series(list(calendar), dtype=object)
    bad_calendar = calendar_series[invalid_date_mask(calendar_series)]
    report["invalid_calendar_dates"] = int(len(bad_calendar))
    report["calendar_dates"] = len(calendar)
    if len(bad_calendar):
        failures.append(
            f"{len(bad_calendar)} invalid calendar dates: {bad_calendar.tolist()[:5]}")
    if list(calendar) != sorted(calendar):
        failures.append("calendar is not sorted ascending")

    report["failures"] = failures
    report["passed"] = not failures
    if failures:
        raise GateFailure(
            "raw OHLC source integrity failed - detect() would be meaningless on "
            "a structurally broken frame:\n  - " + "\n  - ".join(failures))
    return report


def audit_broker_source(frame):
    """Source-integrity audit of the RAW broker harvest, before any conversion.

    Normalization must not run on incoherent inputs: converting units or
    deriving an average price from fields that do not add up would launder a
    data defect into a plausible-looking feature.

    NaN is checked EXPLICITLY rather than inferred from comparisons. `NaN < 0`
    is False and `NaN != x` is True, so a purely comparison-based audit would
    wave non-finite values straight through.

    Sign conventions: blot/slot/bval/sval are gross buy/sell quantities and can
    never be negative. nlot/nval are NET and legitimately can be. Value and lots
    must agree about whether a side traded at all -- in BOTH directions, since
    positive lots with no value is exactly as incoherent as value with no lots.
    """
    report = {"rows": int(len(frame))}
    failures = []

    non_finite = {}
    for col in REQUIRED_BROKER_NUMERIC:
        count = int((~np.isfinite(frame[col].to_numpy(dtype=float))).sum())
        non_finite[col] = count
        if count:
            failures.append(f"{count} non-finite (NaN/inf) values in required field {col}")
    report["non_finite"] = non_finite

    report["duplicate_date_ticker_broker"] = int(
        frame.duplicated(["date", "ticker", "broker"]).sum())
    if report["duplicate_date_ticker_broker"]:
        failures.append(
            f"{report['duplicate_date_ticker_broker']} duplicate (date,ticker,broker) rows")

    bad_date = invalid_date_mask(frame["date"])
    report["invalid_dates"] = int(bad_date.sum())
    if report["invalid_dates"]:
        examples = frame.loc[bad_date, "date"].astype(str).unique()[:5].tolist()
        failures.append(
            f"{report['invalid_dates']} invalid/impossible dates, e.g. {examples}")

    missing_ids = frame["ticker"].isna() | (frame["ticker"].astype(str).str.strip() == "") \
        | frame["broker"].isna() | (frame["broker"].astype(str).str.strip() == "")
    report["missing_identifiers"] = int(missing_ids.sum())
    if report["missing_identifiers"]:
        failures.append(f"{report['missing_identifiers']} rows with missing ticker/broker")

    for col in ("blot", "slot", "bval", "sval"):
        count = int((frame[col] < 0).sum())
        report[f"negative_{col}"] = count
        if count:
            failures.append(f"{count} rows with negative {col} (gross field cannot be negative)")

    nlot_gap = (frame["nlot"] - (frame["blot"] - frame["slot"])).abs()
    report["nlot_inconsistent"] = int((nlot_gap > NLOT_TOLERANCE).sum())
    report["nlot_max_gap"] = float(nlot_gap.max()) if len(frame) else 0.0
    if report["nlot_inconsistent"]:
        failures.append(f"{report['nlot_inconsistent']} rows where nlot != blot - slot")

    nval_gap = (frame["nval"] - (frame["bval"] - frame["sval"])).abs()
    report["nval_inconsistent"] = int((nval_gap > NVAL_TOLERANCE_RUPIAH).sum())
    report["nval_max_gap_rupiah"] = float(nval_gap.max()) if len(frame) else 0.0
    report["nval_tolerance_rupiah"] = NVAL_TOLERANCE_RUPIAH
    if report["nval_inconsistent"]:
        failures.append(
            f"{report['nval_inconsistent']} rows where nval != bval - sval beyond "
            f"{NVAL_TOLERANCE_RUPIAH} rupiah")

    coherence = {
        "bval_positive_blot_nonpositive": int(((frame["bval"] > 0) & (frame["blot"] <= 0)).sum()),
        "sval_positive_slot_nonpositive": int(((frame["sval"] > 0) & (frame["slot"] <= 0)).sum()),
        "blot_positive_bval_nonpositive": int(((frame["blot"] > 0) & (frame["bval"] <= 0)).sum()),
        "slot_positive_sval_nonpositive": int(((frame["slot"] > 0) & (frame["sval"] <= 0)).sum()),
    }
    report.update(coherence)
    for name, count in coherence.items():
        if count:
            failures.append(f"{count} rows failing value/lot coherence: {name}")

    report["failures"] = failures
    report["passed"] = not failures
    return report


# ── harvest ingestion ─────────────────────────

def load_full_harvest(path=OHLC_PARQUET):
    """Load the ENTIRE frozen price harvest plus its trading calendar.

    Returns the full cross-section deliberately. Two things need it:

      - detect()'s cross_ticker_dup, which compares OHLCV ACROSS tickers and is
        blinded by any prior filtering;
      - the trading calendar, which must be every date the market traded, so
        add_forward_returns() can tell a genuine suspension apart from a date
        nobody traded.
    """
    frame = pd.read_parquet(path)
    calendar = sorted(frame["date"].unique())
    return frame, calendar, file_fingerprint(path)


def price_coverage(full_frame, universe, fingerprint, calendar):
    """Which approved names the price harvest carries, asserted against the freeze."""
    present = set(full_frame["ticker"].unique())
    covered = sorted(set(universe) & present)
    missing = sorted(set(universe) - present)

    coverage = {
        "fingerprint": fingerprint,
        "harvest_rows_total": int(len(full_frame)),
        "harvest_tickers_total": int(full_frame["ticker"].nunique()),
        "calendar_dates": len(calendar),
        "calendar_first": calendar[0] if calendar else None,
        "calendar_last": calendar[-1] if calendar else None,
        "price_covered_n": len(covered),
        "price_missing": missing,
    }

    failures = []
    if len(covered) != EXPECTED_PRICE_COVERED:
        failures.append(f"{len(covered)} tickers with prices, expected {EXPECTED_PRICE_COVERED}")
    if tuple(missing) != EXPECTED_PRICE_MISSING:
        failures.append(
            f"price-missing set is {missing}, expected exactly {list(EXPECTED_PRICE_MISSING)}")
    if failures:
        raise GateFailure(
            "price coverage no longer matches the frozen snapshot - stop rather "
            "than silently proceed:\n  - " + "\n  - ".join(failures))
    return covered, coverage


def load_frozen_broker(universe, path=BROKER_PARQUET):
    """Audit the raw broker harvest, then normalize into #1E's unit convention.

    The harvest stores plain rupiah with an explicit buy/sell split
    (blot/bval/slot/sval). #1E's frozen feature code reads value columns in
    BILLIONS and recovers lots from (value, average price), so this converts once
    and derives the average execution price implied by (value, lots).

    Nothing is invented: where a side has zero lots its implied average price is
    undefined and left NaN. That is faithful, but it has a consequence for
    feature fidelity that audit_net_lot_recovery() measures rather than assumes.

    Broker coverage is frozen separately from price coverage, because a broker
    harvest that quietly covered fewer names would shrink variants B/C/D while
    leaving A untouched -- making the variants incomparable without anything
    looking wrong.

    Raises before normalizing if the source audit fails.
    """
    frame = pd.read_parquet(path)
    present = set(frame["ticker"].unique())
    covered = sorted(set(universe) & present)
    missing = sorted(set(universe) - present)

    coverage = {
        "fingerprint": file_fingerprint(path),
        "harvest_rows_total": int(len(frame)),
        "harvest_tickers_total": int(frame["ticker"].nunique()),
        "broker_covered_n": len(covered),
        "broker_missing": missing,
    }
    failures = []
    if len(covered) != EXPECTED_BROKER_COVERED:
        failures.append(
            f"{len(covered)} tickers with broker flow, expected {EXPECTED_BROKER_COVERED}")
    if tuple(missing) != EXPECTED_BROKER_MISSING:
        failures.append(
            f"broker-missing set is {missing}, expected exactly {list(EXPECTED_BROKER_MISSING)}")
    if failures:
        raise GateFailure(
            "broker coverage no longer matches the frozen snapshot - variants "
            "B/C/D would silently shrink:\n  - " + "\n  - ".join(failures))

    rows = frame[frame["ticker"].isin(covered)].copy()
    source_audit = audit_broker_source(rows)
    if not source_audit["passed"]:
        raise GateFailure(
            "broker source integrity audit failed - refusing to normalize "
            "incoherent inputs:\n  - " + "\n  - ".join(source_audit["failures"]))

    rows = rows.rename(columns={"broker": "broker_code"})
    rows["netval"] = rows["nval"] / RUPIAH_PER_BILLION
    rows["bval_b"] = rows["bval"] / RUPIAH_PER_BILLION
    rows["sval_b"] = rows["sval"] / RUPIAH_PER_BILLION

    buy_shares = rows["blot"] * SHARES_PER_LOT
    sell_shares = rows["slot"] * SHARES_PER_LOT
    rows["bavg"] = np.where(buy_shares > 0, rows["bval"] / buy_shares, np.nan)
    rows["savg"] = np.where(sell_shares > 0, rows["sval"] / sell_shares, np.nan)
    rows = rows.drop(columns=["bval", "sval"]).rename(
        columns={"bval_b": "bval", "sval_b": "sval"})

    keep = ["date", "ticker", "broker_code", "nlot", "blot", "slot",
            "netval", "bval", "sval", "bavg", "savg"]
    out = rows[keep].sort_values(["ticker", "date", "broker_code"]).reset_index(drop=True)
    return out, source_audit, coverage


def audit_net_lot_recovery(broker, panel, sample=None):
    """Measure how faithfully #1E's frozen feature code recovers source lots.

    This corrects a claim that would otherwise be wrong. #1E's
    _historical_net_lots() takes its exact buy/sell path ONLY when bval, sval,
    bavg and savg are all non-null AND both averages are > 0. A ONE-SIDED broker
    row -- bought but never sold that day, or the reverse -- has an undefined
    average on the empty side, so it falls through to the netval/close fallback
    and does NOT reproduce the source nlot exactly.

    Gate A does not fix that: changing _historical_net_lots would alter frozen
    #1E behaviour. It measures it instead, so the fidelity limit travels with
    the experiment as a stated caveat rather than an unexamined assumption.

    One-sided means exactly one side traded -- an XOR. A row where NEITHER side
    traded is zero-sided and is counted separately; folding it into "one-sided"
    would misdescribe a broker that simply did not trade that name that day.

    The frozen function is imported and actually executed here -- not
    paraphrased -- so what is measured is precisely what Gate B will compute.
    """
    from ml_v2_experiment_1 import _historical_net_lots

    merged = broker.merge(panel[["date", "ticker", "close"]],
                          on=["date", "ticker"], how="inner")
    if sample is not None and len(merged) > sample:
        merged = merged.sample(n=sample, random_state=17).reset_index(drop=True)

    recovered = _historical_net_lots(merged)
    live_mask = (
        recovered[["bval", "sval", "bavg", "savg"]].notna().all(axis=1)
        & recovered["bavg"].gt(0) & recovered["savg"].gt(0)
    )
    bought = recovered["blot"] > 0
    sold = recovered["slot"] > 0
    one_sided = bought ^ sold
    zero_sided = ~bought & ~sold
    gap = (recovered["net_lots"] - recovered["nlot"]).abs()

    exact_gap = gap[live_mask]
    fallback_gap = gap[~live_mask]
    return {
        "rows_audited": int(len(recovered)),
        "exact_split_path_rows": int(live_mask.sum()),
        "fallback_path_rows": int((~live_mask).sum()),
        "fallback_share": round(float((~live_mask).mean()), 6) if len(recovered) else 0.0,
        "one_sided_rows": int(one_sided.sum()),
        "one_sided_and_fallback": int((one_sided & ~live_mask).sum()),
        "zero_sided_rows": int(zero_sided.sum()),
        "zero_sided_and_fallback": int((zero_sided & ~live_mask).sum()),
        "exact_path_max_lot_gap": float(exact_gap.max()) if len(exact_gap) else 0.0,
        "exact_path_median_lot_gap": float(exact_gap.median()) if len(exact_gap) else 0.0,
        "fallback_path_max_lot_gap": float(fallback_gap.max()) if len(fallback_gap) else 0.0,
        "fallback_path_median_lot_gap": float(fallback_gap.median()) if len(fallback_gap) else 0.0,
        "caveat": "one-sided and zero-sided rows take the netval/close fallback in "
                  "frozen #1E code and do not reproduce source nlot exactly; this is "
                  "a Gate B feature-fidelity caveat, not a defect introduced here",
    }


def broker_provenance(broker, db_path=os.path.join(HERE, "neobdm.db")):
    """Bound what the harvest's broker values can honestly be said to prove.

    Two DIFFERENT claims are at stake and must not be conflated:

      VALUE FIDELITY   - do the harvest's numbers match independently observed
                         ones? Checkable on the window where neobdm.db holds
                         live-scraped rows, and that is what this measures.

      PIT AVAILABILITY - were those numbers obtainable BEFORE the EOD(T) decision
                         cutoff they are used for? NOT checkable here.
                         broker_flow carries no observation timestamp, only a
                         trade date, so nothing in this data says when a row
                         became visible. Same-day post-session availability
                         remains an ASSUMPTION, in every window.

    Value reconciliation must never be read as evidence of the second claim.
    Matching numbers say the harvest recorded the same trades; they say nothing
    about publication timing.
    """
    import sqlite3

    ledger = {
        "harvest_first": broker["date"].min() if len(broker) else None,
        "harvest_last": broker["date"].max() if len(broker) else None,
        "value_reconciled_from": VALUE_RECONCILED_FROM,
        "pit_availability": "ASSUMED, NOT VERIFIED in any window - broker_flow "
                            "carries no observation timestamp, so availability "
                            "before the EOD(T) cutoff is an assumption",
    }
    if not os.path.exists(db_path):
        ledger["status"] = "neobdm.db absent - no value cross-check possible"
        return ledger

    conn = sqlite3.connect(db_path)
    try:
        live = pd.read_sql(
            "SELECT date, ticker, broker_code, bval, sval FROM broker_flow "
            "WHERE bval IS NOT NULL", conn)
    finally:
        conn.close()

    if live.empty:
        ledger["status"] = "no live-scraped rows in neobdm.db"
        return ledger

    overlap = broker[(broker["date"] >= live["date"].min())
                     & (broker["date"] <= live["date"].max())]
    merged = live.merge(overlap, on=["date", "ticker", "broker_code"],
                        how="inner", suffixes=("_db", "_hv"))

    # The two windows must not share a boundary date, or a reader sees the same
    # day claimed as both reconciled and unverified. The unverified window ends
    # on the last harvest date STRICTLY BEFORE reconciliation begins.
    earlier = broker.loc[broker["date"] < VALUE_RECONCILED_FROM, "date"]
    unverified_last = earlier.max() if len(earlier) else None

    ledger.update({
        "live_rows": int(len(live)),
        "live_window": [live["date"].min(), live["date"].max()],
        "matched_rows": int(len(merged)),
        "VALUE_RECONCILED_WINDOW": [VALUE_RECONCILED_FROM, ledger["harvest_last"]],
        "HISTORICAL_UNVERIFIED_WINDOW": [ledger["harvest_first"], unverified_last],
    })
    if merged.empty:
        ledger["status"] = "no overlapping (date,ticker,broker) rows"
        return ledger

    # neobdm.db rounds billions to 1 decimal place, so exact equality is the
    # wrong test. Agreement means the harvest value rounds to the stored one --
    # which is also the proof that the rupiah -> billions conversion is right.
    for col in ("bval", "sval"):
        db_value = merged[f"{col}_db"]
        harvest_rounded = merged[f"{col}_hv"].round(1)
        agree = np.isclose(harvest_rounded, db_value, atol=0.05, equal_nan=True)
        ledger[f"{col}_agreement_rate"] = round(float(agree.mean()), 6)
        ledger[f"{col}_disagreements"] = int((~agree).sum())

    ledger["status"] = "VALUE fidelity reconciled on the live window only"
    return ledger


# ── validated panel construction ──────────────

def build_validated_panel(full_harvest, universe, calendar, horizons=HORIZONS,
                          lags=LAGS):
    """detect() on the FULL harvest, then narrow to the approved universe.

    Order matters and is the point of this function. cross_ticker_dup compares
    OHLCV across tickers, so it must see the whole cross-section: an approved
    ticker cloned against a name OUTSIDE the universe looks unique the moment
    the other half of the pair is filtered away, and the contamination sails
    through. So the detector runs first, on everything, and only then are the
    suspect keys and the model panel restricted to the approved names.

    Out-of-universe tickers contribute detection evidence and calendar dates.
    They never enter the returned panel.

    Returns (panel, flagged_full, universe_rows).
    """
    full = full_harvest.sort_values(["ticker", "date"]).reset_index(drop=True)
    flagged = detect(full)
    quarantined = set(zip(flagged.loc[flagged["suspect"], "date"],
                          flagged.loc[flagged["suspect"], "ticker"]))

    universe_rows = full[full["ticker"].isin(universe)].reset_index(drop=True)
    keep = [(d, t) not in quarantined
            for d, t in zip(universe_rows["date"], universe_rows["ticker"])]
    clean = universe_rows[keep].reset_index(drop=True)

    panel = add_forward_returns(clean, calendar, horizons=horizons, open_anchored=True)
    if lags:
        panel = add_lagged_returns(panel, calendar, lags=lags)
    return panel, flagged, universe_rows


def open_anchor_diagnostics(panel, calendar, horizons=HORIZONS):
    """Separate open-anchor rejections from contiguity/end-of-window attrition.

    A raw count of NaN fwd_oo_* says how much was lost but not why, and the
    causes mean very different things: a failed open anchor is a data-quality
    rejection (a fabricated or out-of-band open), while a contiguity break is a
    calendar fact (a suspension, or simply running off the end of the panel).

    The open-anchor half reuses price_audit._open_anchor_valid directly -- the
    real contract, not a paraphrase. The contiguity half is plain position
    arithmetic on the same unfiltered date axis add_forward_returns uses. The
    remainder is attributed by elimination to the close-step band chain rather
    than re-deriving that chain here, which would be exactly the divergent copy
    this module exists to avoid.
    """
    frame = panel.sort_values(["ticker", "date"]).reset_index(drop=True)
    group = frame.groupby("ticker")
    open_valid = _open_anchor_valid(frame, group).fillna(False).astype(bool)

    positions = {d: i for i, d in enumerate(calendar)}
    pos = frame["date"].map(positions)
    pos_group = pos.groupby(frame["ticker"])

    report = {
        "rows": int(len(frame)),
        "rows_with_invalid_open_anchor": int((~open_valid).sum()),
        "rows_with_valid_open_anchor": int(open_valid.sum()),
        "per_horizon": {},
    }

    entry_ok = open_valid.groupby(frame["ticker"]).shift(-1).fillna(False).astype(bool)
    for h in horizons:
        exit_ok = open_valid.groupby(frame["ticker"]).shift(-(1 + h)).fillna(False).astype(bool)
        contiguous = ((pos_group.shift(-(1 + h)) - pos) == (1 + h)).fillna(False).astype(bool)
        window_exists = pos_group.shift(-(1 + h)).notna()
        invalid = frame[f"fwd_oo_{h}"].isna()

        report["per_horizon"][f"fwd_oo_{h}"] = {
            "non_executable": int(invalid.sum()),
            "entry_open_invalid": int((invalid & ~entry_ok).sum()),
            "exit_open_invalid": int((invalid & ~exit_ok).sum()),
            "window_past_end_of_series": int((invalid & ~window_exists).sum()),
            "non_contiguous_calendar": int((invalid & window_exists & ~contiguous).sum()),
            "close_step_or_other_by_elimination": int(
                (invalid & entry_ok & exit_ok & contiguous).sum()),
        }
    return report


def integrity_checks(universe_rows, panel, flagged_full, universe, horizons=HORIZONS):
    """Post-quarantine structural report. Assertion happens in assert_structural."""
    report = {}

    in_universe = flagged_full["ticker"].isin(universe)
    report["detector_counts_full_harvest"] = {
        name: int(flagged_full[name].sum())
        for name in ("limit_violation", "cross_ticker_dup", "series_break")
    }
    report["detector_counts_universe"] = {
        name: int((flagged_full[name] & in_universe).sum())
        for name in ("limit_violation", "cross_ticker_dup", "series_break")
    }
    report["suspect_rows_full_harvest"] = int(flagged_full["suspect"].sum())
    report["suspect_rows_universe"] = int((flagged_full["suspect"] & in_universe).sum())

    # How many universe rows were quarantined ONLY because their clone lived
    # outside the universe -- the contamination a pre-filtered detector misses.
    dup_only_outside = 0
    dups = flagged_full[flagged_full["cross_ticker_dup"]]
    if len(dups):
        key_cols = ["date"] + OHLCV
        universe_flag = dups["ticker"].isin(universe)
        grouped = dups.groupby(key_cols)["ticker"].transform("nunique")
        inside_count = universe_flag.groupby(
            [dups[c] for c in key_cols]).transform("sum")
        dup_only_outside = int(((universe_flag) & (inside_count == 1) & (grouped > 1)).sum())
    report["universe_rows_dup_against_outside_only"] = dup_only_outside

    scoped = flagged_full[in_universe]
    reason_parts = []
    for name in ("limit_violation", "cross_ticker_dup", "series_break"):
        reason_parts.append(np.where(scoped[name], name, ""))
    reasons = pd.Series(
        ["+".join(p for p in parts if p) for parts in zip(*reason_parts)],
        index=scoped.index, dtype=object) if len(scoped) else pd.Series(dtype=object)
    report["suspect_reason_breakdown_universe"] = (
        reasons[scoped["suspect"].to_numpy()].value_counts().to_dict()
        if len(scoped) else {})

    report["rows_in"] = int(len(universe_rows))
    report["rows_after_quarantine"] = int(len(panel))
    report["tickers_after_quarantine"] = int(panel["ticker"].nunique())
    report["out_of_universe_rows_in_panel"] = int((~panel["ticker"].isin(universe)).sum())

    report["duplicate_date_ticker"] = int(panel.duplicated(["date", "ticker"]).sum())
    report["blank_dates"] = int(invalid_date_mask(panel["date"]).sum())
    report["nonpositive_close"] = int((panel["close"] <= 0).sum())
    report["open_outside_high_low"] = int(
        ((panel["open"] < panel["low"]) | (panel["open"] > panel["high"])).sum())

    # Cross-ticker contamination must not survive quarantine WITHIN the panel.
    surviving_dup = panel.duplicated(["date"] + OHLCV, keep=False) \
        & panel[OHLCV].notna().all(axis=1)
    report["surviving_cross_ticker_dup"] = int(surviving_dup.sum())

    ordered = panel.sort_values(["ticker", "date"])
    report["dates_monotonic_per_ticker"] = bool(
        ordered.groupby("ticker")["date"].apply(lambda s: s.is_monotonic_increasing).all())

    # Ticker-boundary integrity: a label may never be computed across two
    # different tickers. Every ticker's FIRST row cannot have a lag, and its
    # LAST row cannot have a forward label.
    first_rows = ordered.groupby("ticker").head(1)
    last_rows = ordered.groupby("ticker").tail(1)
    report["first_row_lag_leak"] = (
        int(first_rows["lag_1"].notna().sum()) if "lag_1" in panel else 0)
    report["last_row_forward_leak"] = int(last_rows[f"fwd_oo_{horizons[0]}"].notna().sum())

    label_cols = [f"fwd_oo_{h}" for h in horizons]
    report["label_coverage"] = {
        col: {
            "valid": int(panel[col].notna().sum()),
            "non_executable": int(panel[col].isna().sum()),
            "valid_pct": round(float(panel[col].notna().mean()), 6),
        }
        for col in label_cols
    }
    # DIAGNOSTIC ONLY, never an invariant. Each horizon validates its OWN exit
    # open, so a longer horizon can legitimately be valid where a shorter one is
    # not -- if H2's particular exit open is out of band while H3's is fine.
    report["valid_count_decreases_with_horizon"] = bool(
        all(panel[label_cols[i]].notna().sum() >= panel[label_cols[i + 1]].notna().sum()
            for i in range(len(label_cols) - 1)))

    band = panel[label_cols].abs().max().max()
    report["max_abs_label"] = None if pd.isna(band) else round(float(band), 6)
    return report


def assert_structural_integrity(report):
    """Hard gate on invariants no downstream code could survive.

    Deliberately excluded: open_outside_high_low and any other open-anchor
    condition. Those rows are ALLOWED to remain -- the contract withholds their
    executable labels instead. Deleting them would be a repair.
    """
    failures = []
    if report["duplicate_date_ticker"]:
        failures.append(f"{report['duplicate_date_ticker']} duplicate (date,ticker) rows")
    if report["blank_dates"]:
        failures.append(f"{report['blank_dates']} blank/unparseable/impossible dates")
    if report["nonpositive_close"]:
        failures.append(f"{report['nonpositive_close']} non-positive closes")
    if report["surviving_cross_ticker_dup"]:
        failures.append(
            f"{report['surviving_cross_ticker_dup']} cross-ticker identical OHLCV rows "
            "survived quarantine")
    if report["out_of_universe_rows_in_panel"]:
        failures.append(
            f"{report['out_of_universe_rows_in_panel']} out-of-universe rows leaked "
            "into the model panel")
    if report["first_row_lag_leak"]:
        failures.append(
            f"{report['first_row_lag_leak']} ticker-boundary lag leak(s): a first row "
            "carries a lag it cannot have")
    if report["last_row_forward_leak"]:
        failures.append(
            f"{report['last_row_forward_leak']} ticker-boundary forward leak(s): a last "
            "row carries a forward label it cannot have")
    if not report["dates_monotonic_per_ticker"]:
        failures.append("dates are not monotonically increasing within every ticker")
    if failures:
        raise GateFailure(
            "post-quarantine structural integrity failed:\n  - " + "\n  - ".join(failures))
    return True


def start_of_history(universe_rows, panel):
    """First-appearance statistics from the RAW source, before quarantine.

    Computed on the raw covered rows on purpose: if a ticker's opening row
    happens to be quarantined, the post-quarantine first date would misreport
    when the SOURCE history actually begins, and a reader would mistake a
    dropped bad bar for a later listing. Both views are reported so the
    difference is visible rather than assumed away.

    Neither view is a listing date. No point-in-time membership data exists in
    this repo, so first appearance is the only proxy available -- and for most
    names it records when COLLECTION started, not when the stock listed.
    """
    raw_firsts = universe_rows.groupby("ticker")["date"].min()
    clean_firsts = panel.groupby("ticker")["date"].min()
    floor = raw_firsts.min()
    shifted = int((clean_firsts > raw_firsts.reindex(clean_firsts.index)).sum())
    return {
        "source_harvest_floor": floor,
        "source_tickers_starting_at_floor": int((raw_firsts == floor).sum()),
        "source_tickers_starting_later": int((raw_firsts > floor).sum()),
        "source_first_date_histogram": raw_firsts.value_counts().sort_index().head(12).to_dict(),
        "post_quarantine_first_differs": shifted,
        "post_quarantine_floor": clean_firsts.min(),
        "bars_per_ticker_min": int(panel.groupby("ticker").size().min()),
        "bars_per_ticker_median": int(panel.groupby("ticker").size().median()),
        "bars_per_ticker_max": int(panel.groupby("ticker").size().max()),
    }


def run_gate(xlsx_path=UNIVERSE_XLSX, refreeze=False, establish_manifest=False,
             net_lot_sample=None):
    """Full Gate A pass. Returns (panel, broker, report)."""
    tickers, universe_audit = resolve_universe(xlsx_path, refreeze=refreeze)

    full_harvest, calendar, ohlc_fingerprint = load_full_harvest()
    covered, price_cov = price_coverage(full_harvest, tickers, ohlc_fingerprint, calendar)
    broker, broker_source, broker_cov = load_frozen_broker(tickers)

    manifest = verify_input_manifest(
        [ohlc_fingerprint, broker_cov["fingerprint"]], establish=establish_manifest)

    raw_audit = audit_raw_ohlc(full_harvest, calendar, scope="full harvest")
    panel, flagged_full, universe_rows = build_validated_panel(
        full_harvest, covered, calendar)
    price_cov["universe_rows"] = int(len(universe_rows))

    integrity = integrity_checks(universe_rows, panel, flagged_full, covered)
    assert_structural_integrity(integrity)

    report = {
        "universe": universe_audit,
        "input_manifest": manifest,
        "price_coverage": price_cov,
        "broker_coverage": broker_cov,
        "raw_ohlc_audit": raw_audit,
        "broker_source_audit": broker_source,
        "integrity": integrity,
        "open_anchor": open_anchor_diagnostics(panel, calendar),
        "history": start_of_history(universe_rows, panel),
        "net_lot_recovery": audit_net_lot_recovery(broker, panel, sample=net_lot_sample),
        "broker_provenance": broker_provenance(broker),
        "broker_rows": int(len(broker)),
        "broker_tickers": int(broker["ticker"].nunique()),
        "broker_codes": int(broker["broker_code"].nunique()),
        "panel_digest": frame_digest(
            panel, ["date", "ticker", "open", "high", "low", "close", "volume"]
            + [f"fwd_oo_{h}" for h in HORIZONS] + ["gap_1"]),
        "broker_digest": frame_digest(
            broker, ["date", "ticker", "broker_code", "netval", "bval", "sval"]),
    }
    return panel, broker, report


def print_report(report):
    universe = report["universe"]
    price_cov = report["price_coverage"]
    broker_cov = report["broker_coverage"]
    raw = report["raw_ohlc_audit"]
    integrity = report["integrity"]
    anchors = report["open_anchor"]
    history = report["history"]
    provenance = report["broker_provenance"]
    source = report["broker_source_audit"]
    lots = report["net_lot_recovery"]

    print("=" * 78)
    print("EXPERIMENT #1F - GATE A: DATA-INTEGRATION VALIDATION")
    print("=" * 78)

    print("\n[1] APPROVED UNIVERSE")
    print(f"  loaded from       : {universe.get('loaded_from', 'frozen artifact')}")
    print(f"  source            : {universe['source_path']}")
    print(f"  sheet / header    : {universe['sheet']!r} / {universe['header_cell']!r}")
    print(f"  distinct tickers  : {universe['distinct_tickers']} (expected {EXPECTED_UNIVERSE_SIZE})")
    print(f"  digest            : {universe['universe_digest']} (accepted {ACCEPTED_UNIVERSE_DIGEST})")
    print(f"  non-string cells  : {universe['non_string_cells'] or 'none (no Excel type coercion)'}")
    print(f"  blank cells       : {universe['blank_cells'] or 'none'}")
    print(f"  malformed         : {universe['malformed_tickers'] or 'none'}")
    print(f"  duplicates        : {universe['duplicates'] or 'none'}")

    print("\n[2] FROZEN INPUT MANIFEST (missing manifest is a FAILURE, never auto-adopted)")
    print(f"  status   : {report['input_manifest']['status']}")
    for name, values in sorted(report["input_manifest"]["inputs"].items()):
        print(f"  {name:<20}: {values['sha256']}")
        print(f"  {'':<20}  {values['size_bytes']} bytes")

    print("\n[3] COVERAGE (price and broker frozen independently)")
    print(f"  full harvest      : {price_cov['harvest_rows_total']} rows / "
          f"{price_cov['harvest_tickers_total']} tickers")
    print(f"  trading calendar  : {price_cov['calendar_dates']} dates "
          f"({price_cov['calendar_first']} -> {price_cov['calendar_last']})")
    print(f"  PRICE covered     : {price_cov['price_covered_n']} / "
          f"{universe['distinct_tickers']} (expected {EXPECTED_PRICE_COVERED})")
    print(f"  PRICE missing     : {price_cov['price_missing']} (excluded, never fabricated)")
    print(f"  BROKER covered    : {broker_cov['broker_covered_n']} / "
          f"{universe['distinct_tickers']} (expected {EXPECTED_BROKER_COVERED})")
    print(f"  BROKER missing    : {broker_cov['broker_missing']}")
    print(f"  universe rows     : {price_cov['universe_rows']}")

    print("\n[4] RAW OHLC SOURCE INTEGRITY (FULL harvest, before detect())")
    print(f"  scope                      : {raw['scope']}")
    print(f"  rows                       : {raw['rows']}")
    print(f"  duplicate (date,ticker)    : {raw['duplicate_date_ticker']}")
    print(f"  invalid/impossible dates   : {raw['invalid_dates']} (strict %Y-%m-%d parse)")
    print(f"  missing ticker             : {raw['missing_ticker']}")
    print(f"  non-finite OHLCV           : {raw['non_finite']}")
    print(f"  non-positive o/h/l/c       : {raw['nonpositive']}")
    print(f"  invalid calendar dates     : {raw['invalid_calendar_dates']} "
          f"of {raw['calendar_dates']}")
    print(f"  passed                     : {raw['passed']}")

    print("\n[5] BROKER SOURCE INTEGRITY (raw rupiah, before normalization)")
    print(f"  rows                          : {source['rows']}")
    print(f"  non-finite required fields    : {source['non_finite']}")
    print(f"  duplicate (date,ticker,broker): {source['duplicate_date_ticker_broker']}")
    print(f"  invalid/impossible dates      : {source['invalid_dates']}")
    print(f"  missing ticker/broker ids     : {source['missing_identifiers']}")
    print(f"  negative gross fields         : blot {source['negative_blot']} | "
          f"slot {source['negative_slot']} | bval {source['negative_bval']} | "
          f"sval {source['negative_sval']}")
    print(f"  nlot != blot-slot             : {source['nlot_inconsistent']} "
          f"(max gap {source['nlot_max_gap']:.6g})")
    print(f"  nval != bval-sval             : {source['nval_inconsistent']} "
          f"(max gap {source['nval_max_gap_rupiah']:.6g} rupiah, tol "
          f"{source['nval_tolerance_rupiah']})")
    print(f"  value/lot coherence           : bval>0&blot<=0 "
          f"{source['bval_positive_blot_nonpositive']} | sval>0&slot<=0 "
          f"{source['sval_positive_slot_nonpositive']} | blot>0&bval<=0 "
          f"{source['blot_positive_bval_nonpositive']} | slot>0&sval<=0 "
          f"{source['slot_positive_sval_nonpositive']}")
    print(f"  passed                        : {source['passed']}")

    print("\n[6] DETECTORS (price_audit.detect on the FULL cross-section)")
    print(f"  {'detector':<20}{'full harvest':>14}{'universe':>11}")
    for name in ("limit_violation", "cross_ticker_dup", "series_break"):
        print(f"  {name:<20}{integrity['detector_counts_full_harvest'][name]:>14}"
              f"{integrity['detector_counts_universe'][name]:>11}")
    print(f"  {'suspect rows':<20}{integrity['suspect_rows_full_harvest']:>14}"
          f"{integrity['suspect_rows_universe']:>11}")
    print(f"  universe rows cloned ONLY against an out-of-universe ticker: "
          f"{integrity['universe_rows_dup_against_outside_only']}")
    print("  (those would have been INVISIBLE had the detector run on the subset)")
    for reason, count in sorted(integrity["suspect_reason_breakdown_universe"].items()):
        print(f"      {reason:<32}: {count}")

    print("\n[7] POST-QUARANTINE STRUCTURAL INTEGRITY (hard gate - all must be 0/True)")
    print(f"  rows in / after quarantine : {integrity['rows_in']} -> "
          f"{integrity['rows_after_quarantine']} "
          f"({integrity['tickers_after_quarantine']} tickers)")
    print(f"  out-of-universe rows leaked: {integrity['out_of_universe_rows_in_panel']}")
    print(f"  duplicate (date,ticker)    : {integrity['duplicate_date_ticker']}")
    print(f"  blank/impossible dates     : {integrity['blank_dates']}")
    print(f"  non-positive close         : {integrity['nonpositive_close']}")
    print(f"  surviving cross-ticker dup : {integrity['surviving_cross_ticker_dup']}")
    print(f"  ticker-boundary lag leak   : {integrity['first_row_lag_leak']}")
    print(f"  ticker-boundary fwd leak   : {integrity['last_row_forward_leak']}")
    print(f"  dates monotonic per ticker : {integrity['dates_monotonic_per_ticker']}")
    print("  --- reported, NOT gated ---")
    print(f"  open outside [low,high]    : {integrity['open_outside_high_low']} "
          f"(anchor contract withholds the label; row is kept)")
    print(f"  max |open-anchored label|  : {integrity['max_abs_label']}")

    print("\n[8] OPEN-ANCHOR REJECTION DIAGNOSTICS (Hn = open(T+1) -> open(T+1+n))")
    print(f"  rows with invalid open anchor : {anchors['rows_with_invalid_open_anchor']} "
          f"of {anchors['rows']}")
    print(f"  {'label':<11}{'NON_EXEC':>10}{'entry_bad':>11}{'exit_bad':>10}"
          f"{'past_end':>10}{'non_contig':>12}{'close/other':>13}")
    for col, stats in anchors["per_horizon"].items():
        print(f"  {col:<11}{stats['non_executable']:>10}{stats['entry_open_invalid']:>11}"
              f"{stats['exit_open_invalid']:>10}{stats['window_past_end_of_series']:>10}"
              f"{stats['non_contiguous_calendar']:>12}"
              f"{stats['close_step_or_other_by_elimination']:>13}")
    print("  (categories overlap by construction; 'close/other' is by elimination)")

    print("\n[9] LABEL COVERAGE")
    print(f"  {'label':<12}{'valid':>10}{'NON_EXEC':>12}{'valid %':>10}")
    for col, stats in integrity["label_coverage"].items():
        print(f"  {col:<12}{stats['valid']:>10}{stats['non_executable']:>12}"
              f"{stats['valid_pct']:>10.4f}")
    print(f"  valid count falls with horizon (DIAGNOSTIC, not required): "
          f"{integrity['valid_count_decreases_with_horizon']}")
    print("  Each horizon validates its OWN exit open, so a longer horizon may")
    print("  legitimately be valid where a shorter one is not.")

    print("\n[10] START-OF-HISTORY (from RAW source, pre-quarantine)")
    print(f"  source harvest floor         : {history['source_harvest_floor']}")
    print(f"  tickers starting at floor    : {history['source_tickers_starting_at_floor']}")
    print(f"  tickers starting later       : {history['source_tickers_starting_later']}")
    print(f"  post-quarantine first differs: {history['post_quarantine_first_differs']} ticker(s)")
    print(f"  bars/ticker min/median/max   : {history['bars_per_ticker_min']} / "
          f"{history['bars_per_ticker_median']} / {history['bars_per_ticker_max']}")
    print("  NOTE: first appearance is a COLLECTION date, not a listing date.")

    print("\n[11] NET-LOT RECOVERY FIDELITY (frozen #1E _historical_net_lots)")
    print(f"  rows audited                 : {lots['rows_audited']}")
    print(f"  exact split path             : {lots['exact_split_path_rows']}")
    print(f"  fallback (netval/close) path : {lots['fallback_path_rows']} "
          f"({lots['fallback_share']:.4%})")
    print(f"  one-sided rows (XOR)         : {lots['one_sided_rows']} "
          f"({lots['one_sided_and_fallback']} take the fallback)")
    print(f"  zero-sided rows (neither)    : {lots['zero_sided_rows']} "
          f"({lots['zero_sided_and_fallback']} take the fallback)")
    print(f"  |recovered-nlot| exact path  : median {lots['exact_path_median_lot_gap']:.6g} "
          f"| max {lots['exact_path_max_lot_gap']:.6g}")
    print(f"  |recovered-nlot| fallback    : median {lots['fallback_path_median_lot_gap']:.6g} "
          f"| max {lots['fallback_path_max_lot_gap']:.6g}")
    print(f"  CAVEAT: {lots['caveat']}")

    print("\n[12] BROKER PROVENANCE LEDGER")
    print(f"  rows / tickers / codes       : {report['broker_rows']} / "
          f"{report['broker_tickers']} / {report['broker_codes']}")
    print(f"  harvest window               : {provenance.get('harvest_first')} -> "
          f"{provenance.get('harvest_last')}")
    if "VALUE_RECONCILED_WINDOW" in provenance:
        window = provenance["VALUE_RECONCILED_WINDOW"]
        unverified = provenance["HISTORICAL_UNVERIFIED_WINDOW"]
        print(f"  HISTORICAL_UNVERIFIED_WINDOW : {unverified[0]} -> {unverified[1]}")
        print(f"  VALUE_RECONCILED_WINDOW      : {window[0]} -> {window[1]}")
        print(f"  matched rows for crosscheck  : {provenance['matched_rows']} "
              f"(of {provenance['live_rows']} live rows)")
        print(f"  bval agreement rate          : {provenance.get('bval_agreement_rate')} "
              f"({provenance.get('bval_disagreements')} disagreements)")
        print(f"  sval agreement rate          : {provenance.get('sval_agreement_rate')} "
              f"({provenance.get('sval_disagreements')} disagreements)")
    print(f"  status                       : {provenance.get('status')}")
    print(f"  PIT availability             : {provenance['pit_availability']}")

    print("\n[13] DIGESTS")
    print(f"  universe : {universe['universe_digest']}")
    print(f"  panel    : {report['panel_digest']}")
    print(f"  broker   : {report['broker_digest']}")

    print("\n[14] STANDING CAVEATS")
    print("  - Current-universe retrospective/survivorship bias is NOT eliminated by")
    print("    listing-proxy filtering. No PIT membership data exists in this repo.")
    print("  - Harvest floor is a COLLECTION date, not a listing date.")
    print("  - Broker VALUE fidelity is reconciled only on the live window; broker")
    print("    PIT AVAILABILITY is assumed and unverified in EVERY window.")
    print("  - One-sided/zero-sided broker rows take #1E's fallback path; see [11].")
    print("  - Snapshot is frozen, not refreshed: OOS ends at the calendar last date.")
    print("=" * 78)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--universe", default=UNIVERSE_XLSX)
    parser.add_argument("--json-out", default=None,
                        help="write the machine-readable report here")
    parser.add_argument("--determinism-check", action="store_true",
                        help="build twice and require identical digests")
    parser.add_argument("--refreeze", action="store_true",
                        help="deliberately re-read the workbook and re-freeze "
                             "(still hard-fails if its content changed)")
    parser.add_argument("--establish-manifest", action="store_true",
                        help="explicitly establish/re-establish the frozen input "
                             "manifest after reviewing the hashes")
    parser.add_argument("--net-lot-sample", type=int, default=None,
                        help="audit net-lot recovery on a seeded sample of N rows")
    args = parser.parse_args()

    panel, broker, report = run_gate(args.universe, refreeze=args.refreeze,
                                     establish_manifest=args.establish_manifest,
                                     net_lot_sample=args.net_lot_sample)
    print_report(report)

    if args.determinism_check:
        panel2, broker2, report2 = run_gate(args.universe,
                                            net_lot_sample=args.net_lot_sample)
        same_panel = report["panel_digest"] == report2["panel_digest"]
        same_broker = report["broker_digest"] == report2["broker_digest"]
        equal_frames = panel.equals(panel2) and broker.equals(broker2)
        if not (same_panel and same_broker and equal_frames):
            raise GateFailure(
                f"Gate A is not deterministic: panel {report['panel_digest']} vs "
                f"{report2['panel_digest']}, broker {report['broker_digest']} vs "
                f"{report2['broker_digest']}, frames_equal={equal_frames}")
        print(f"\nDeterminism check: PASS (panel {report['panel_digest']}, "
              f"broker {report['broker_digest']}, frames identical)")

    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as handle:
            json.dump(report, handle, indent=2, sort_keys=True, default=str)
            handle.write("\n")
        print(f"\nMachine-readable report: {args.json_out}")


if __name__ == "__main__":
    main()
